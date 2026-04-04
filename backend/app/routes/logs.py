from email.mime import message
from gettext import find
from logging import critical, log
from unittest import result
import uuid
import re
import os
import csv
import json
import asyncio
from fastapi.concurrency import run_in_threadpool
from datetime import datetime, time, timezone
from typing import Optional
from reportlab.lib import styles # pyright: ignore[reportMissingModuleSource]
from sqlalchemy import or_
from fastapi import APIRouter, Form, UploadFile, File, Depends, BackgroundTasks
from sqlalchemy.orm import Session
from fastapi import Query
from openpyxl import load_workbook

from app.database import get_db, SessionLocal
from app.models.threat_log import ThreatLog
from app.models.user_activity import UserActivity
from app.models.incident import Incident
from app.models.asset import Asset
from app.security import require_role
from app.websocket.manager import manager

from app.services.alert_service import create_alert
from app.services.detection_engine import detect_port_scan, run_detection_engine
from app.services.response_engine import automated_response
from app.services.audit_service import log_action
from app.services.correlation_engine import correlate_alert
from app.services.automation_engine import auto_response
from app.models import alert
from fastapi import Body

from app.routes import incidents

router = APIRouter(prefix="/logs", tags=["Logs"])

DEFAULT_SOC_TARGET = "192.168.1.10"

MITRE_MAP = {
    "scan": ("Discovery", "T1046"),
    "probe": ("Discovery", "T1046"),
    "brute": ("Credential Access", "T1110"),
    "login failed": ("Credential Access", "T1110"),
    "malware": ("Execution", "T1059"),
    "exploit": ("Execution", "T1059"),
    "callback": ("Command and Control", "T1071"),
    "c2": ("Command and Control", "T1071"),
    "suspicious": ("Persistence", "T1543"),
    "service": ("Persistence", "T1543"),
    "url": ("Command and Control", "T1071"),   # 🔥 ADD THIS
    "http": ("Command and Control", "T1071"),  # 🔥 ADD THIS
    "https": ("Command and Control", "T1071"), # 🔥 ADD THIS
}

progress_store = {}
# 🔥 EMAIL COOLDOWN STORE
email_cooldown = {}

def parse_timestamp(row):
    """
    Extract timestamp from Excel/CSV/log row
    """

    for val in row:
        if not val:
            continue

        # ✅ Excel datetime object
        if isinstance(val, datetime):
            return val

        val_str = str(val).strip()

        # ✅ ISO format (2026-03-21T14:40:00)
        try:
            return datetime.fromisoformat(val_str)
        except:
            pass

        # ✅ Common formats (Excel / logs)
        for fmt in [
            "%d/%m/%Y %I:%M:%S %p",   # 21/03/2026 02:41:46 PM
            "%d-%m-%Y %H:%M:%S",      # 21-03-2026 14:41:46
            "%Y-%m-%d %H:%M:%S",      # 2026-03-21 14:41:46
        ]:
            try:
                return datetime.strptime(val_str, fmt)
            except:
                continue

    return None

@router.get("")
def get_logs(
    page: int = Query(1),
    limit: int = Query(100),
    db: Session = Depends(get_db),
    user=Depends(require_role("ADMIN", "ANALYST", "VIEWER")),
):
    offset = (page - 1) * limit

    total = db.query(ThreatLog).filter(
        ThreatLog.user_email == user["sub"]
    ).count()

    logs = (
        db.query(ThreatLog)
        .filter(ThreatLog.user_email == user["sub"])
        .order_by(ThreatLog.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    items = []

    for log in logs:
        items.append(
            {
                "src_ip": log.source_ip,
                "dst_ip": log.destination_ip,
                "src_port": log.source_port,
                "dst_port": log.destination_port,
                "protocol": log.protocol,
                "severity": log.severity,

                # 🔥 FIX START
                "risk_score": log.risk_score if log.risk_score is not None else 0,
                # 🔥 FIX END

                "status": log.status, 
                "threat": log.message,
                "created_at": log.created_at.isoformat() if log.created_at else None,  # ✅ FIX HERE # type: ignore
                "mitre_tactic": log.mitre_tactic,
                "mitre_technique": log.mitre_technique,
            }
        )

    return {
        "items": items,
        "total": total
    }


@router.post("/parse")
async def parse_logs(
    background_tasks: BackgroundTasks,
    raw_text: str = Form(""),
    file: Optional[UploadFile] = File(None),
    user=Depends(require_role("ADMIN", "ANALYST", "VIEWER")),
    db: Session = Depends(get_db),
):

    if not file and not raw_text:
        return {"error": "No file uploaded or raw logs provided"}

    db.query(ThreatLog).delete()
    db.query(Incident).delete()
    

    UPLOAD_DIR = "uploads"
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    filename = file.filename.lower() if file and file.filename else "raw_input.log"
    file_path = f"{UPLOAD_DIR}/{filename}"

    if file:
        with open(file_path, "wb") as f:
            f.write(await file.read())
    else:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(raw_text)

    background_tasks.add_task(
        process_logs,
        file_path,
        filename,
        user["sub"]
    )

    return {"message": "Log processing started in background"}


def process_logs(file_path, filename, username):

    db = SessionLocal()

    logs_to_add = []
    alerts_to_stream = []
    
    # 🚀 PERFORMANCE BOOST (ADD HERE)
    existing_ips = set(
        ip for (ip,) in db.query(Asset.ip).all()
    )

    try:

        rows_cache = []
        total_rows = 0
        processed = 0

        if filename.endswith(".xlsx"):
            workbook = load_workbook(data_only=True, read_only=True, filename=file_path)
            sheet = workbook.active

            rows = list(sheet.iter_rows(values_only=True)) # type: ignore

            if not rows:
                return

            # ✅ HEADER DETECTION
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]

            print("HEADERS:", headers)  # optional debug

            for row in rows[1:]:
                rows_cache.append((headers, row))

        elif filename.endswith(".csv"):
            with open(file_path, encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                rows = list(reader)

                if rows:
                    headers = [str(h).lower().strip() for h in rows[0]]

                    for row in rows[1:]:
                        rows_cache.append((headers, row))

        elif filename.endswith(".json"):
            with open(file_path, encoding="utf-8", errors="ignore") as f:
                data = json.load(f)
            for entry in data:
                rows_cache.append([
                    entry.get("time"),
                    entry.get("severity"),
                    None,
                    None,
                    entry.get("message"),
                    entry.get("src_ip"),
                    entry.get("dst_ip"),
                    entry.get("protocol")
                ])

        elif filename.endswith(".log") or filename.endswith(".txt"):
            with open(file_path, encoding="utf-8", errors="ignore") as f:
                for line in f:
                    rows_cache.append([None, "LOW", None, None, line.strip()])

        else:
            print("Unsupported file format")
            return
        
        total_rows = len(rows_cache)

        progress_store[username] = {
            "total": total_rows,
            "processed": 0
        }

        def parse_row(data):
    
            try:

                # ✅ HANDLE BOTH TYPES
                if isinstance(data, tuple):
                    headers, row = data

                    row_dict = {}
                    for i, key in enumerate(headers):
                        if key and i < len(row):
                            row_dict[key] = row[i]

                    def find(keywords):
                        for k, v in row_dict.items():
                            for keyword in keywords:
                                if keyword in k:
                                    return v
                        return None

                    source_ip = find(["src", "source"])
                    destination_ip = find(["dst", "destination"])
                    source_port = find(["src port", "source port"]) or "0"
                    destination_port = find(["dst port", "destination port"]) or "0"
                    protocol = (find(["protocol"]) or "UNKNOWN").upper()
                    message = find(["message", "threat", "log"]) or "No message"
                    event_time = find(["time", "date"]) or datetime.now(timezone.utc)
                    file_severity = find(["severity", "level"])
                    
                    if not source_ip:
                        return

                else:
                    # 🔥 SMART PARSING (NON-EXCEL FILES)
                    row = data
                    row_text = " ".join([str(x) for x in row if x])

                    # 🔥 IP DETECTION
                    ip_matches = re.findall(r"\b\d{1,3}(?:\.\d{1,3}){3}\b", row_text)
                    source_ip = ip_matches[0] if ip_matches else "0.0.0.0"
                    destination_ip = ip_matches[1] if len(ip_matches) > 1 else DEFAULT_SOC_TARGET

                    # 🔥 PORT DETECTION
                    port_matches = re.findall(r":(\d{1,5})", row_text)
                    if len(port_matches) >= 2:
                        source_port = port_matches[0]
                        destination_port = port_matches[1]
                    elif len(port_matches) == 1:
                        source_port = port_matches[0]
                        destination_port = "0"
                    else:
                        source_port = "0"
                        destination_port = "0"

                    # 🔥 PROTOCOL DETECTION
                    protocol = "Unknown"
                    if re.search(r"\btcp\b", row_text, re.IGNORECASE):
                        protocol = "TCP"
                    elif re.search(r"\budp\b", row_text, re.IGNORECASE):
                        protocol = "UDP"
                    elif re.search(r"\bicmp\b", row_text, re.IGNORECASE):
                        protocol = "ICMP"
                    elif re.search(r"\bhttp\b", row_text, re.IGNORECASE):
                        protocol = "HTTP"
                    elif re.search(r"\bhttps\b", row_text, re.IGNORECASE):
                        protocol = "HTTPS"

                    # 🔥 MESSAGE
                    message = row_text
                    
                    file_severity = None  # ✅ FIX: avoid undefined variable

                    # 🔥 TIME DETECTION
                    event_time = parse_timestamp(row) or datetime.now(timezone.utc)

                # ✅ FIX TIME
                if isinstance(event_time, str):
                    event_time = parse_timestamp([event_time]) or datetime.now(timezone.utc)

                row_text = str(message)
                row_text_lower = row_text.lower()

                # ================= KEEP YOUR EXISTING LOGIC =================

                # AUTO CREATE ASSET
                try:
                    # 🚀 FAST ASSET CHECK (REPLACE HERE)
                    if source_ip not in existing_ips:
                        existing_ips.add(source_ip)

                        new_asset = Asset(
                            id=str(uuid.uuid4()),
                            ip=source_ip,
                            hostname="Unknown",
                            owner="Auto-Discovered",
                            criticality="LOW"
                        )
                        db.add(new_asset)

                except Exception as e:
                    print("Asset creation failed:", e)

                # PROTOCOL NORMALIZATION
                if file_severity and str(file_severity).strip():
                    severity = str(file_severity).strip().upper()
                else:
                    # SEVERITY DETECTION
                    severity = "LOW"

                    if re.search(r"(critical|ransomware|data breach|root access)", row_text_lower):
                        severity = "CRITICAL"

                    elif re.search(r"(attack|brute force|exploit|malware|unauthorized)", row_text_lower):
                        severity = "HIGH"

                    elif re.search(r"(scan|probe|suspicious|recon)", row_text_lower):
                        severity = "MEDIUM"

                    elif re.search(r"(failed|invalid|denied)", row_text_lower):
                        severity = "LOW"
                        
                    # 🔥 DETECTION ENGINE FIX
                    try:
                        result = detect_port_scan(db, source_ip)

                        if result:
                            print("DETECTION TRIGGERED:", result)

                    except Exception as e:
                        print("Detection failed:", e)

                risk_map = {
                    "CRITICAL": 95,
                    "HIGH": 75,
                    "MEDIUM": 50,
                    "LOW": 20,
                }

                risk_score = risk_map.get(severity, 10)

                # 🔥 ADD THIS BLOCK
                status = "SAFE"

                if severity in ["CRITICAL", "HIGH"]:
                    status = "THREAT"
                elif severity == "MEDIUM":
                    status = "SUSPICIOUS"

                    try:
                        create_alert(
                            db=db,
                            source_ip=source_ip,
                            severity=severity,
                            message=row_text,
                            risk_score=risk_score,
                            classification=row_text
                        )

                        # 🔥 FIX: LOG = INCIDENT (NO GROUPING)
                        new_incident = Incident(
                            id=str(uuid.uuid4()),
                            source_ip=source_ip,
                            severity=severity or "LOW",
                            status="OPEN",
                            created_at=event_time or datetime.now(timezone.utc),
                            alert_count=1
                        )

                        db.add(new_incident)

                    except Exception as e:
                            print("Alert/Incident failed:", e)

                # MITRE
                mitre_tactic = None
                mitre_technique = None
                for key, value in MITRE_MAP.items():
                    if key in row_text_lower:
                        mitre_tactic, mitre_technique = value
                        break
                    
                # 🔥 FALLBACK (IMPORTANT)
                if not mitre_tactic:
                    mitre_tactic = "Command and Control"
                    mitre_technique = "T1071"

                # ✅ SAFETY FIX (ADD HERE)
                protocol = protocol if protocol else "UNKNOWN"
                
                # SAVE LOG
                log = ThreatLog(
                    id=str(uuid.uuid4()),
                    source_ip=source_ip,
                    destination_ip=destination_ip,
                    source_port=str(source_port),
                    destination_port=str(destination_port),
                    protocol=protocol,
                    severity=severity,
                    risk_score=risk_score,
                    classification=row_text,
                    message=row_text,
                    status=status,
                    created_at=event_time,
                    mitre_tactic=mitre_tactic,
                    mitre_technique=mitre_technique,
                    user_email=username
                )

                logs_to_add.append(log)
                
                if len(logs_to_add) >= 2000:
                    db.add_all(logs_to_add)
                    db.commit()
                    logs_to_add.clear()
                
                # 🔥 RUN DETECTION ENGINE
                try:
                    detection_result = run_detection_engine(db, log)

                    # ✅ ADD THIS BLOCK (VERY IMPORTANT)
                    if detection_result and processed % 200 == 0: # type: ignore
                        print("🚨 DETECTED:", detection_result)

                        alert_data = {
                            "source_ip": log.source_ip,
                            "severity": log.severity,
                            "message": f"{detection_result} detected",
                            "risk_score": log.risk_score,
                        }

                        alerts_to_stream.append(alert_data)
                    
                except Exception as e:
                    print("Detection engine error:", e)

            except Exception as e:
                print("ROW ERROR:", e)
                
        for row in rows_cache:
            parse_row(row)
            processed += 1
            
            progress_store[username]["processed"] = processed
            
            if processed % 500 == 0 or processed == total_rows:
                try:
                    asyncio.run(manager.broadcast({
                        "type": "PROGRESS_UPDATE",
                        "processed": processed,
                        "total": total_rows
                    }))
                except:
                    pass

        if logs_to_add:
            db.add_all(logs_to_add)

        log_action(
            db,
            "LOG_IMPORT",
            username,
            details=f"{len(logs_to_add)} logs imported",
            page="logs"
        )
        
        

        db.commit()
        
        progress_store[username] = {
            "processed": progress_store[username]["total"],
            "total": progress_store[username]["total"]
        }

        # ✅ FIXED (SAFE ASYNC EXECUTION)
        async def safe_broadcast():
            for alert in alerts_to_stream[:100]:
                try:
                    # ✅ SEND ALERT
                    await manager.broadcast({
                        "type": "NEW_ALERT",
                        "severity": alert["severity"],
                        "source_ip": alert["source_ip"],
                        "message": alert["message"],
                        "risk_score": alert["risk_score"],
                    })

                    # ✅ SEND PROGRESS SEPARATELY
                    await manager.broadcast({
                        "type": "PROGRESS_UPDATE",
                        "processed": progress_store[username]["processed"],
                        "total": progress_store[username]["total"]
                    })
                except Exception as e:
                    print("Broadcast failed:", e)

        try:
            loop = asyncio.get_running_loop()
            asyncio.create_task(safe_broadcast())  # ✅ BEST FIX
        except RuntimeError:
            asyncio.run(safe_broadcast())
        

    finally:
        db.close()

@router.get("/search")
def search_logs(
        query: str = Query(default=""),
        source_ip: str = Query(default=""),
        destination_ip: str = Query(default=""),
        severity: str = Query(default=""),
        protocol: str = Query(default=""),
        page: int = 1,
        limit: int = 50,
        db: Session = Depends(get_db),
        user=Depends(require_role("ADMIN", "ANALYST", "VIEWER")),
    ):
    
    print("QUERY RECEIVED:", repr(query))

    offset = (page - 1) * limit

    q = db.query(ThreatLog).filter(
        ThreatLog.user_email == user["sub"]
    )
    
    # ✅ Prevent empty search
    if not query and not source_ip and not destination_ip:
        return {"items": []}

    # ✅ Main query search    
    if query and query.strip():
        ip = query.strip()
        
        print("🔍 Searching IP:", ip)

        q = q.filter(
            or_(
                ThreatLog.source_ip.ilike(f"%{ip}%"),
                ThreatLog.destination_ip.ilike(f"%{ip}%")
            )
        )

    elif source_ip:
        ip = source_ip.strip()
        q = q.filter(
            or_(
                ThreatLog.source_ip == ip,
                ThreatLog.destination_ip == ip
            )
        )
    
    if destination_ip:
        q = q.filter(ThreatLog.destination_ip == destination_ip.strip())

    if severity:
        q = q.filter(ThreatLog.severity == severity.upper())

    if protocol:
        q = q.filter(ThreatLog.protocol == protocol.upper())

    logs = (
    q.order_by(ThreatLog.created_at.desc())
    .offset(offset)
    .limit(limit)
    .all()
    )

    items = []

    for log in logs:
        items.append({
            "src_ip": log.source_ip or "N/A",
            "dst_ip": log.destination_ip or "N/A",
            "src_port": log.source_port or "N/A",
            "dst_port": log.destination_port or "N/A",
            "protocol": log.protocol or "UNKNOWN",
            "severity": log.severity or "LOW",
            "severity": log.severity or "LOW",

            # 🔥 FIX START (PASTE HERE)
            "risk_score": log.risk_score if log.risk_score is not None else 0,
            # 🔥 FIX END

            "status": log.status,
            "threat": log.message or "N/A",      # ✅ ADD
            "status": log.status, 
            "threat": log.message or "N/A",
            "created_at": log.created_at.isoformat() if log.created_at else None, # type: ignore
            "mitre_tactic": log.mitre_tactic or "N/A",
            "mitre_technique": log.mitre_technique or "N/A",
        })

    return {"items": items}

@router.get("/hunt")
def threat_hunt(
    ip: str,
    db: Session = Depends(get_db),
    user=Depends(require_role("ADMIN", "ANALYST", "VIEWER")),
):
    from sqlalchemy import or_

    logs = db.query(ThreatLog).filter(
        ThreatLog.user_email == user["sub"],
        or_(
            ThreatLog.source_ip == ip,
            ThreatLog.destination_ip == ip
        )
    ).order_by(ThreatLog.created_at.desc()).limit(200).all()

    if not logs:
        return {"items": []}

    items = []

    for log in logs:
        items.append({
            "src_ip": log.source_ip,
            "dst_ip": log.destination_ip,
            "protocol": log.protocol,
            "severity": log.severity,
            "risk_score": log.risk_score if log.risk_score is not None else 0,     # ✅ ADD THIS
            "status": log.status, 
            "threat": log.message,
            "time": log.created_at.isoformat() if log.created_at else None # type: ignore
        })

    return {"items": items}

# ================= PDF DOWNLOAD =================

from fastapi.responses import StreamingResponse
from reportlab.platypus import Flowable, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.lib.styles import getSampleStyleSheet # pyright: ignore[reportMissingModuleSource]
from reportlab.graphics.shapes import Drawing # pyright: ignore[reportMissingModuleSource]
from reportlab.graphics.charts.barcharts import VerticalBarChart # pyright: ignore[reportMissingModuleSource]
import io
import smtplib
from email.message import EmailMessage
from datetime import datetime
from collections import Counter

# ✅ GEOIP SAFE IMPORT
try:
    import geoip2.database # pyright: ignore[reportMissingImports]
except ImportError:
    geoip2 = None

# ================= AI SCORE =================
def ai_threat_score(severity, message):

    score = 0

    if severity == "CRITICAL":
        score += 90
    elif severity == "HIGH":
        score += 70
    elif severity == "MEDIUM":
        score += 50
    else:
        score += 20

    msg = (message or "").lower()

    if "malware" in msg:
        score += 10
    if "attack" in msg:
        score += 10
    if "unauthorized" in msg:
        score += 5
    if "scan" in msg:
        score += 5

    return min(score, 100)


# ================= AI CLASSIFIER =================
def classify_threat(text: str):
    t = (text or "").lower()

    if "sql" in t:
        return "SQL Injection"
    elif "brute" in t or "login failed" in t:
        return "Brute Force"
    elif "scan" in t:
        return "Port Scan"
    elif "malware" in t:
        return "Malware"
    return "General Threat"


@router.post("/download")
def download_logs_pdf(
    logs: list = Body(...),
    company: str = Body(default="AegisCyber SOC"),
    analyst: str = Body(default="SOC Analyst"),
    user_email: Optional[str] = Body(default=None),
    db: Session = Depends(get_db),
    user = Depends(require_role("ADMIN", "ANALYST", "VIEWER")),  # ✅ ADD THIS
    
):

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=(600, 800),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=30
    )
    
    # 🔥 ENTERPRISE HEADER (ADD HERE)
    def draw_header(canvas, doc):
        canvas.saveState()

        # Logo
        logo_path = os.path.join(os.getcwd(), "app", "assets", "logo.png")
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, 30, 750, width=35, height=35)

        # Title
        canvas.setFont("Helvetica-Bold", 14)
        canvas.setFillColor(colors.white)
        canvas.drawString(80, 770, f"{company} - SOC Report")

        # Date
        canvas.setFont("Helvetica", 9)
        canvas.drawRightString(570, 770, datetime.now().strftime("%Y-%m-%d %H:%M"))

        canvas.restoreState()

    def add_footer(canvas, doc):
        canvas.saveState()

        # Dark footer strip
        canvas.setFillColorRGB(0.1, 0.1, 0.1)
        canvas.rect(0, 0, doc.pagesize[0], 25, fill=1)

        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica", 8)

        canvas.drawString(30, 10, "QuickSOC | Confidential Security Report")
        canvas.drawRightString(570, 10, f"Page {doc.page}")

        canvas.restoreState()
    styles = getSampleStyleSheet()
    styles["Normal"].textColor = colors.white
    styles["Heading2"].textColor = colors.lime
    styles["Title"].textColor = colors.aqua

    elements = []
    
    from reportlab.platypus import Spacer # type: ignore

    elements.append(Spacer(1, 10))  # 🔥 small top padding (fix gap)

    # ================= BACKGROUND =================
    from reportlab.platypus import Flowable # type: ignore

    # 🔥 DARK BACKGROUND (HACKER STYLE)
    from reportlab.platypus import Flowable # type: ignore

    class Background(Flowable):
        def draw(self):
            self.canv.setFillColorRGB(0.05, 0.07, 0.12)  # dark blue-black
            self.canv.rect(0, 0, 600, 800, fill=1)

    elements.append(Background())

    # ================= HEADER =================
    elements.append(Paragraph(
        f"<para align='center'><font size=26 color='#00ffcc'><b>{company}</b></font></para>",
        styles["Title"]
    ))
    logo_path = os.path.join(os.getcwd(), "app", "assets", "logo.png")

    if os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=40, height=40)
            logo.hAlign = 'CENTER'
            elements.append(logo)
        except:
            print("Logo failed")
    elements.append(Spacer(1, 2))
    elements.append(Paragraph(
        "<para align='center'><font size=10 color='#38bdf8'>AI Threat Intelligence • Real-Time Detection • Cyber Defense System</font></para>",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 6))
    

    elements.append(Paragraph(
        "<para align='center'><font size=14><b>Security Operations Center Report</b></font></para>",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 6))

    elements.append(Paragraph(f"<b>Generated On:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Prepared By:</b> {analyst}", styles["Normal"]))
    elements.append(Paragraph("<b>System:</b> SOC Threat Monitoring Platform", styles["Normal"]))

    elements.append(Spacer(1, 10))

    # ================= INCIDENT FILTER =================
    incidents = [
        log for log in logs
        if log.get("severity", "").upper() in ["MEDIUM", "HIGH", "CRITICAL"]
    ]
    
    severity_counts = Counter(
    log.get("severity", "LOW").upper() for log in incidents
    )
    # 🔥 FIX: DEFINE email_summary
    email_summary = {
        "CRITICAL": severity_counts.get("CRITICAL", 0),
        "HIGH": severity_counts.get("HIGH", 0),
        "MEDIUM": severity_counts.get("MEDIUM", 0),
        "LOW": severity_counts.get("LOW", 0),
    }

    if not incidents:
        elements.append(Paragraph("No security incidents found.", styles["Normal"]))
    
    
    # ================= SUMMARY =================
    elements.append(Paragraph(
        "<font color='#00ff00'><b>[ SYSTEM ANALYSIS REPORT ]</b></font>",
        styles["Heading2"]
    ))
    summary_box = Table([[
        f"Logs: {len(logs)}",
        f"Incidents: {len(incidents)}",
        f"Critical: {severity_counts.get('CRITICAL',0)}",
        f"High: {severity_counts.get('HIGH',0)}",
        f"Medium: {severity_counts.get('MEDIUM',0)}"
    ]])

    summary_box.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.black),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.lime),
        ("GRID", (0,0), (-1,-1), 0.5, colors.green),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))
    
    # ================= EXECUTIVE SUMMARY =================
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<b>📄 Executive Summary</b>", styles["Heading2"]))

    summary_text = f"""
    Total {len(incidents)} security incidents detected.
    Majority are {severity_counts.get('MEDIUM',0)} medium-level threats.
    No critical breaches observed.
    System is stable but requires monitoring.
    """

    elements.append(Paragraph(summary_text, styles["Normal"]))

    elements.append(summary_box)
    elements.append(Spacer(1, 8))
    elements.append(Table([[""]], colWidths=[550], rowHeights=[2],
        style=[("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#22c55e"))]
    ))  # green separator line
    
    # 🔥 SEVERITY COLOR BAR (ADD HERE)
    severity_bar = Table([[
        f"CRITICAL ({severity_counts.get('CRITICAL',0)})",
        f"HIGH ({severity_counts.get('HIGH',0)})",
        f"MEDIUM ({severity_counts.get('MEDIUM',0)})",
        f"LOW ({severity_counts.get('LOW',0)})"
    ]])

    severity_bar.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,0), colors.red),
        ('BACKGROUND', (1,0), (1,0), colors.orange),
        ('BACKGROUND', (2,0), (2,0), colors.yellow),
        ('BACKGROUND', (3,0), (3,0), colors.green),

        ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
    ]))

    elements.append(Spacer(1, 10))
    elements.append(severity_bar)
    
    # ================= TOP ATTACKERS (MOVE HERE) =================
    ip_counts = Counter()
    for log in incidents:
        ip = log.get("src_ip")
        if ip:
            ip_counts[ip] += 1

    top_ips = ip_counts.most_common(5)

    # ================= AI ANALYSIS =================
    ai_text = "No major threats detected."

    if severity_counts.get("CRITICAL", 0) > 0:
        ai_text = "Critical threats detected. Immediate action required."
    elif severity_counts.get("HIGH", 0) > 0:
        ai_text = "High-risk activity observed. Investigation recommended."
    elif severity_counts.get("MEDIUM", 0) > 0:
        ai_text = "Moderate suspicious activity detected. Monitor closely."
    else:
        ai_text = "System operating normally."

    elements.append(Spacer(1, 10))
    elements.append(Paragraph(
        "<font color='#00ffff'><b>🧠 AI Threat Analysis</b></font>",
        styles["Heading2"]
    ))

    primary_threat = max(severity_counts.keys(), key=lambda k: severity_counts[k]) if severity_counts else "None"

    detailed_ai = f"""
    The system analyzed {len(logs)} logs and identified {len(incidents)} threats.

    Top attacker IP: {top_ips[0][0] if top_ips else "N/A"}  
    Most common severity: {primary_threat}

    ⚠ Risk Insight:
    Repeated traffic patterns suggest automated attack behavior.

    ✅ Recommended Actions:
    - Block high-frequency IPs
    - Enable IDS/IPS rules
    - Monitor unusual traffic spikes
    """

    elements.append(Paragraph(detailed_ai, styles["Normal"]))
    elements.append(Spacer(1, 8))
    elements.append(Table([[""]], colWidths=[550], rowHeights=[2],
        style=[("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#06b6d4"))]
    ))  # cyan separator line
    
    # ================= CHART =================
    elements.append(Paragraph("<b>📈 Threat Distribution</b>", styles["Heading2"]))
    chart_data = [
        severity_counts.get("CRITICAL", 0),
        severity_counts.get("HIGH", 0),
        severity_counts.get("MEDIUM", 0),
    ]

    drawing = Drawing(400, 200)
    chart = VerticalBarChart()

    chart.x = 50
    chart.y = 30
    chart.height = 125
    chart.width = 300
    chart.data = [[
        max(1, severity_counts.get("CRITICAL", 0)),
        max(1, severity_counts.get("HIGH", 0)),
        max(1, severity_counts.get("MEDIUM", 0))
    ]]
    chart.valueAxis.valueMax = max(chart_data) + 10
    chart.valueAxis.valueStep = max(1, int(max(chart_data) / 5))
    chart.bars[0].fillColor = colors.HexColor("#38bdf8")  # clean blue (enterprise look)
    chart.bars[1].fillColor = colors.HexColor("#f97316")  # orange (enterprise look)
    chart.bars[2].fillColor = colors.HexColor("#eab308")  # yellow (enterprise look)

    chart.categoryAxis.categoryNames = ["Critical", "High", "Medium"]

    drawing.add(chart)
    elements.append(drawing)

    elements.append(Spacer(1, 20))
    
    elements.append(Paragraph("<b>🕒 Attack Timeline</b>", styles["Heading2"]))

    timeline_counts = {}

    for log in incidents[:50]:
        time = log.get("created_at", "")
        if time:
            hour = time[:13]  # YYYY-MM-DD HH
            timeline_counts[hour] = timeline_counts.get(hour, 0) + 1

    timeline_data = list(timeline_counts.values())[:10]
    timeline_labels = list(timeline_counts.keys())[:10]

    if timeline_data:
        drawing2 = Drawing(400, 200)
        chart2 = VerticalBarChart()

        chart2.x = 50
        chart2.y = 30
        chart2.height = 125
        chart2.width = 300
        chart2.data = [timeline_data]
        chart2.categoryAxis.categoryNames = timeline_labels

        drawing2.add(chart2)
        elements.append(drawing2)

    elements.append(Spacer(1, 20))
    
    # ================= TOP ATTACKERS =================

    elements.append(Paragraph("<b>🎯 Top Attacker IPs</b>", styles["Heading2"]))

    # 🔥 ADD THIS (PASS TO EMAIL)
    top_attackers_list = [
        {"ip": ip, "count": count}
        for ip, count in top_ips
    ]

    ip_table_data = [["IP Address", "Attack Count"]]

    for ip, count in top_ips:
        ip_table_data.append([ip, str(count)])

    ip_table = Table(ip_table_data)

    ip_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.black),
        ("TEXTCOLOR", (0,0), (-1,0), colors.lime),
        ("GRID", (0,0), (-1,-1), 0.3, colors.green),
    ]))

    elements.append(ip_table)
    elements.append(Spacer(1, 15))
    elements.append(Table([[""]], colWidths=[550], rowHeights=[2],
    style=[("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#f97316"))]
    ))  # orange separator line
    
    if top_ips:
        elements.append(Paragraph(
            f"<b>🔥 Most Active Attacker:</b> {top_ips[0][0]} ({top_ips[0][1]} attacks)",
            styles["Normal"]
        ))      

    # ================= GEOIP =================
    elements.append(Paragraph("<b>🌐 Geo Intelligence</b>", styles["Heading2"]))
    geo_summary = {}

    try:
        if geoip2:
            reader = geoip2.database.Reader("GeoLite2-City.mmdb")
        else:
            reader = None

        for log in incidents:
            ip = log.get("src_ip")
            if not reader:
                continue

            try:
                response = reader.city(ip)
                country = response.country.name
                geo_summary[country] = geo_summary.get(country, 0) + 1
            except:
                pass
    except:
        pass

    if geo_summary:
        geo_text = "<b>Geo Distribution:</b><br/>"
        for k, v in geo_summary.items():
            geo_text += f"{k}: {v}<br/>"

        elements.append(Paragraph(geo_text, styles["Normal"]))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("<b>🔥 Risk Heat Distribution</b>", styles["Heading2"]))

    risk_counts = {"LOW": 0, "MEDIUM": 0, "HIGH": 0, "CRITICAL": 0}

    for log in incidents:
        sev = log.get("severity", "LOW").upper()
        if sev in risk_counts:
            risk_counts[sev] += 1

    heat_data = [[
        "LOW", "MEDIUM", "HIGH", "CRITICAL"
    ], [
        risk_counts["LOW"],
        risk_counts["MEDIUM"],
        risk_counts["HIGH"],
        risk_counts["CRITICAL"]
    ]]

    heat_table = Table(heat_data)

    heat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.black),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("BACKGROUND", (0, 1), (0, 1), colors.green),
        ("BACKGROUND", (1, 1), (1, 1), colors.yellow),
        ("BACKGROUND", (2, 1), (2, 1), colors.orange),
        ("BACKGROUND", (3, 1), (3, 1), colors.red),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))

    elements.append(heat_table)
    elements.append(Spacer(1, 20))

    # ================= TABLE =================
    elements.append(Paragraph("<b>🚨 Incident Details</b>", styles["Heading2"]))
    elements.append(Paragraph(
        "<font color='#facc15'>Showing Top 25 Incidents</font>",
        styles["Normal"]
    ))
    data = [[
        "Source IP",
        "Severity",
        "AI Type",
        "AI Score",
        "Risk",
        "Message"
    ]]

    for log in incidents[:50]:
        ai_score = ai_threat_score(
            log.get("severity"),
            log.get("threat")
        )

        data.append([
            log.get("src_ip", "N/A"),                log.get("severity", "N/A"),
            classify_threat(log.get("threat")),
            str(ai_score),
            str(log.get("risk") or log.get("risk_score") or "N/A"),
            log.get("threat", "N/A")[:40]
        ])

    table = Table(data[:26], repeatRows=1)

    # ✅ MUST MATCH 6 COLUMNS
    table._argW = [70, 60, 80, 60, 50, 200]# type: ignore # control width

    table.setStyle(TableStyle([
        ("TEXTCOLOR", (3,1), (3,-1), colors.cyan),
        ("BACKGROUND", (0, 0), (-1, 0), colors.black),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.lime),

        ("GRID", (0, 0), (-1, -1), 0.2, colors.green),

        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.white),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)

    doc.build(
        elements,
        onFirstPage=lambda c, d: (draw_header(c, d), add_footer(c, d)),
        onLaterPages=lambda c, d: (draw_header(c, d), add_footer(c, d))
    )
    print("🔥 PDF GENERATED SUCCESSFULLY")
    buffer.seek(0)
    
    user_email = user.get("sub") if user else None   # ✅ ADD THIS
    print("👤 Logged in user email:", user_email)
    
    print("📧 STARTING EMAIL PROCESS...")   # ✅ ADD 
    
    import time
    now = time.time()

    # 🔥 PREVENT SPAM (5 MIN COOLDOWN)
    if user_email in email_cooldown and now - email_cooldown[user_email] < 300:
        print("⏸ Skipping email (cooldown active)")
    else:
        email_cooldown[user_email] = now
    # 🔥 AUTO EMAIL SEND
    try:
        fixed_admin_email = "soc.platform11@gmail.com"

        print("📧 Sending to ADMIN:", fixed_admin_email)

        # ✅ send to SOC mailbox
        if user_email not in email_cooldown or now - email_cooldown[user_email] > 300:

            send_email_with_pdf(
                buffer.getvalue(),
                fixed_admin_email,
                email_summary,
                top_attackers_list
            )

        if user_email and user_email != fixed_admin_email:
            print("📧 Sending to USER:", user_email)
        
            send_email_with_pdf(
                buffer.getvalue(),
                user_email,
                email_summary,
                top_attackers_list
            )
        else:
            print("⚠️ No valid user email found")

    except Exception as e:
        print("❌ Email error:", e)
    
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=SOC_Report.pdf"}
    )
    
    # ================= EMAIL FUNCTION =================
def send_email_with_pdf(pdf_bytes, recipient_email, summary=None, top_attackers=None):

    critical = summary.get("CRITICAL", 0) if summary else 0
    high = summary.get("HIGH", 0) if summary else 0
    medium = summary.get("MEDIUM", 0) if summary else 0
    low = summary.get("LOW", 0) if summary else 0

    # 🔥 AI SUMMARY
    if critical > 0:
        ai_summary = "Critical threats detected. Immediate action required."
    elif high > 0:
        ai_summary = "High risk activity observed."
    elif medium > 0:
        ai_summary = "Moderate suspicious activity detected."
    else:
        ai_summary = "System operating normally."

    print("📧 Sending email to:", recipient_email)

    msg = EmailMessage()
    msg["Subject"] = "SOC Threat Report"
    msg["From"] = os.getenv("EMAIL_USER") or "soc.platform11@gmail.com"
    msg["To"] = recipient_email

    # 🔥 TOP ATTACKERS HTML
    top_attackers_html = "<ul>"

    if top_attackers:
        for attacker in top_attackers:
            top_attackers_html += f"<li>{attacker['ip']} ({attacker['count']} attacks)</li>"
    else:
        top_attackers_html += "<li>No attacker data</li>"

    top_attackers_html += "</ul>"

    html_content = f"""
    <html>
    <body style="font-family: Arial; background:#0f172a; color:white; padding:20px;">
        <div style="display:flex; align-items:center; gap:10px;">
            <h2 style="color:#22c55e; margin:0;">AegisCyber SOC</h2>
        </div>

        <p style="color:#94a3b8;">Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>

        <h3 style="color:#facc15;">AI Threat Analysis</h3>
        <p>{ai_summary}</p>

        <h3 style="color:#38bdf8;">Threat Summary</h3>
        <div style="display:flex; gap:10px; flex-wrap:wrap; margin-top:10px;">
            <div style="background:red; color:white; padding:10px;">CRITICAL: {critical}</div>
            <div style="background:orange; color:white; padding:10px;">HIGH: {high}</div>
            <div style="background:yellow; color:black; padding:10px;">MEDIUM: {medium}</div>
            <div style="background:green; color:white; padding:10px;">LOW: {low}</div>
        </div>

        <h3 style="margin-top:20px; color:#f97316;">Top Attackers</h3>
        {top_attackers_html}

        <div style="margin-top:25px;">
            <a href="http://localhost:3000/logs"
               style="background:#22c55e; color:black; padding:12px 20px; text-decoration:none; border-radius:6px;">
               View in Dashboard
            </a>
        </div>

        <p style="margin-top:20px;">PDF report attached.</p>
    </body>
    </html>
    """

    msg.set_content("SOC Alert Notification - Please view the HTML email or attached PDF.")
    msg.add_alternative(html_content, subtype="html")

    if pdf_bytes:
        msg.add_attachment(
            pdf_bytes,
            maintype="application",
            subtype="pdf",
            filename="SOC_Report.pdf"
        )

    try:
        print("📧 Sending email to:", recipient_email)
        print("🔐 EMAIL_USER:", os.getenv("EMAIL_USER"))
        print("🔐 EMAIL_PASS exists:", bool(os.getenv("EMAIL_PASS")))
        
        print("Connecting to SMTP...")
        
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(
                os.getenv("EMAIL_USER") or "",
                os.getenv("EMAIL_PASS") or ""
            )
            
            print("✅ LOGIN SUCCESS")
            
            smtp.send_message(msg)
            print("✅ Email sent successfully to", recipient_email)

    except Exception as e:
        print("❌ Email failed for", recipient_email, ":", e)

@router.get("/progress")
def get_progress(
    user=Depends(require_role("ADMIN", "ANALYST", "VIEWER"))
):
    username = user["sub"]

    progress = progress_store.get(username)

    # ✅ FIX: ensure always valid response
    if not progress:
        return {
            "processed": 0,
            "total": 0,
            "percentage": 0
        }

    processed = progress.get("processed", 0)
    total = progress.get("total", 0)

    # ✅ EXTRA: send percentage (helps frontend)
    percentage = (processed / total * 100) if total > 0 else 0

    return {
        "processed": processed,
        "total": total,
        "percentage": round(percentage, 2)
    }